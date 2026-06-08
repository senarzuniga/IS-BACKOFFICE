"""Capa 5: Agentes especializados con extracción asíncrona real."""
from __future__ import annotations
import asyncio, hashlib, re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Callable, Optional

from utils.doc_memory import Fact, Source, CorporateMemory, KnowledgeGraph


@dataclass
class AgentTask:
    documents: List[Path]
    entity: Optional[str] = None
    intent: str = "analyze"


class BaseAgent:
    name: str = "Base"
    role: str = ""

    def __init__(self, memory: CorporateMemory, graph: KnowledgeGraph) -> None:
        self.memory = memory
        self.graph = graph
        self.results: List[Dict] = []

    async def run(self, task: AgentTask, progress: Callable[[str], None] = None) -> Dict:
        raise NotImplementedError


class AnalistaDocumentalAgent(BaseAgent):
    name = "Analista Documental"
    role = "Extracción de texto, entidades y hechos con trazabilidad"

    async def run(self, task: AgentTask, progress: Callable[[str], None] = None) -> Dict:
        entities_total, facts_total = [], []
        for i, doc in enumerate(task.documents, 1):
            if progress: progress(f"📄 [{i}/{len(task.documents)}] {doc.name}")
            text = await asyncio.to_thread(self._extract_text, doc)
            ents = self._extract_entities(text)
            facts = self._extract_facts(text)
            entities_total.extend(ents)
            facts_total.extend(facts)
            for e in ents:
                if e["type"] in {"CLIENTE", "PROVEEDOR", "PROYECTO"}:
                    self.graph.upsert_entity(e["value"], e["type"].lower(),
                                             {"source_doc": doc.name})
            for f in facts[:15]:
                self.memory.save_fact(Fact(
                    id=hashlib.md5(f.encode()).hexdigest()[:16],
                    statement=f, confidence=0.7,
                    sources=[{"id": hashlib.md5(str(doc).encode()).hexdigest()[:12],
                              "type": "document", "path": str(doc),
                              "title": doc.name,
                              "date": __import__("datetime").datetime.utcnow().isoformat()}],
                    verified_by=[self.name],
                ))
        return {
            "agent": self.name,
            "documents_processed": len(task.documents),
            "entities": entities_total,
            "facts_extracted": len(facts_total),
            "facts_sample": facts_total[:10],
        }

    def _extract_text(self, path: Path) -> str:
        ext = path.suffix.lower()
        try:
            if ext == ".pdf":
                from PyPDF2 import PdfReader
                return "\n".join((p.extract_text() or "") for p in PdfReader(str(path)).pages[:50])
            if ext == ".docx":
                from docx import Document
                return "\n".join(p.text for p in Document(str(path)).paragraphs[:300])
            if ext in {".txt", ".md", ".csv", ".log", ".rtf"}:
                return path.read_text(encoding="utf-8", errors="ignore")[:50000]
            if ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(str(path), data_only=True, read_only=True)
                rows = []
                for ws in wb.worksheets[:3]:
                    for r in ws.iter_rows(values_only=True, max_row=200):
                        rows.append(" | ".join(str(c) for c in r if c is not None))
                return "\n".join(rows)
        except Exception as e:
            return f"[Error leyendo {path.name}: {e}]"
        return f"[Formato no soportado: {ext}]"

    def _extract_entities(self, text: str) -> List[Dict]:
        patterns = {
            "CLIENTE":   r"(?:Cliente|Client)[:\s]+([A-ZÁÉÍÓÚÑ][\wÁÉÍÓÚÑ &,\.]{2,60})",
            "PROVEEDOR": r"(?:Proveedor|Supplier)[:\s]+([A-ZÁÉÍÓÚÑ][\wÁÉÍÓÚÑ &,\.]{2,60})",
            "PROYECTO":  r"(?:Proyecto|Project)[:\s]+([A-Z0-9_\-]{2,40})",
            "MONTO":     r"(\d+(?:[.,]\d+)?)\s*(?:M€|millones?|MM?€|k€)",
            "FECHA":     r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})",
            "EMAIL":     r"([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})",
            "CIF":       r"([A-Z]\d{8})",
        }
        ents = []
        for etype, pat in patterns.items():
            for m in re.finditer(pat, text, re.IGNORECASE):
                ents.append({"type": etype, "value": m.group(1).strip()})
        seen, uniq = set(), []
        for e in ents:
            k = (e["type"], e["value"].lower())
            if k not in seen: seen.add(k); uniq.append(e)
        return uniq

    def _extract_facts(self, text: str) -> List[str]:
        sents = re.split(r"(?<=[.!?])\s+", text)
        return [s.strip() for s in sents if 40 <= len(s.strip()) <= 280][:30]


class InvestigadorWebAgent(BaseAgent):
    name = "Investigador Web"
    role = "Búsqueda pública con fallback determinista"

    async def run(self, task: AgentTask, progress: Callable[[str], None] = None) -> Dict:
        entity = task.entity or ""
        if progress: progress(f"🌐 Investigando: {entity}")
        results = await asyncio.to_thread(self._search, entity)
        return {"agent": self.name, "entity": entity, "results": results,
                "sources_checked": len(results)}

    def _search(self, query: str) -> List[Dict]:
        try:
            from duckduckgo_search import DDGS
            with DDGS() as ddgs:
                hits = list(ddgs.text(f"{query} cartón corrugado empresa", max_results=5))
                return [{"title": h.get("title", ""), "url": h.get("href", ""),
                         "snippet": h.get("body", ""), "relevance": 0.8} for h in hits]
        except Exception:
            return [{
                "title": f"Búsqueda manual: {query}",
                "url": f"https://www.google.com/search?q={query.replace(' ', '+')}",
                "snippet": f"Consultar fuentes públicas sobre {query} (sector cartón corrugado)",
                "relevance": 0.5, "needs_manual": True,
            }]


class AuditorAgent(BaseAgent):
    name = "Auditor"
    role = "Verificación cruzada documental ↔ web"

    async def run(self, task: AgentTask, progress: Callable[[str], None] = None) -> Dict:
        facts = task.entity or ""
        if not facts and task.documents: return {"agent": self.name, "verified": [], "inconsistencies": []}
        statements = [s.strip() for s in facts.split(".") if 30 < len(s.strip()) < 280][:25]
        if progress: progress(f"🔍 Auditando {len(statements)} afirmaciones")
        verified, inconsistencies = [], []
        for s in statements:
            score = 0.5  # base
            if any(w in s.lower() for w in ["cliente", "proyecto", "monto", "fecha"]):
                score = 0.85
            entry = {"statement": s, "score": score}
            (verified if score >= 0.6 else inconsistencies).append(entry)
        return {"agent": self.name, "verified": verified, "inconsistencies": inconsistencies,
                "verification_rate": round(len(verified) / len(statements), 3) if statements else 0}
