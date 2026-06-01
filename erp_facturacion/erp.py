import sqlite3
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font
import os

DB = os.path.join(os.path.dirname(__file__), "database.db")
INVOICE_DIR = os.path.join(os.path.dirname(__file__), "invoices")

# =========================
# BASE DE DATOS
# =========================
def init_db():
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        nif TEXT,
        address TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER,
        date TEXT,
        services REAL,
        expenses REAL,
        base REAL,
        iva REAL,
        irpf REAL,
        total REAL,
        net REAL
    )
    """)

    conn.commit()
    conn.close()

# =========================
# CREAR CLIENTE
# =========================
def add_client(name, nif, address):
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    c.execute("INSERT INTO clients (name, nif, address) VALUES (?, ?, ?)",
              (name, nif, address))

    conn.commit()
    conn.close()

# =========================
# OBTENER CLIENTE
# =========================
def get_client(client_id):
    conn = sqlite3.connect(DB)
    c = conn.cursor()

    c.execute("SELECT * FROM clients WHERE id=?", (client_id,))
    client = c.fetchone()

    conn.close()
    return client

# =========================
# GENERAR FACTURA (CÁLCULO)
# =========================
def create_invoice(client_id, services, expenses, irpf_rate=0.07):
    iva_rate = 0.21
    base = services + expenses
    iva = base * iva_rate
    irpf = base * irpf_rate
    total = base + iva - irpf
    net = base - irpf

    conn = sqlite3.connect(DB)
    c = conn.cursor()

    c.execute("""
    INSERT INTO invoices 
    (client_id, date, services, expenses, base, iva, irpf, total, net)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        client_id,
        str(date.today()),
        services,
        expenses,
        base,
        iva,
        irpf,
        total,
        net
    ))

    invoice_id = c.lastrowid

    conn.commit()
    conn.close()

    generate_excel(invoice_id, client_id, services, expenses, base, iva, irpf, total, net)

    return invoice_id

# =========================
# EXPORTAR A EXCEL
# =========================
def generate_excel(invoice_id, client_id, services, expenses, base, iva, irpf, total, net):
    client = get_client(client_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Factura"

    ws["A1"] = "FACTURA PROFESIONAL"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "Cliente"
    ws["B3"] = client[1]

    ws["A4"] = "NIF"
    ws["B4"] = client[2]

    ws["A6"] = "Servicios"
    ws["B6"] = services

    ws["A7"] = "Gastos"
    ws["B7"] = expenses

    ws["A8"] = "Base imponible"
    ws["B8"] = base

    ws["A9"] = "IVA 21%"
    ws["B9"] = iva

    ws["A10"] = "IRPF 7%"
    ws["B10"] = irpf

    ws["A11"] = "TOTAL FACTURA"
    ws["B11"] = total

    ws["A12"] = "NETO (sin IVA)"
    ws["B12"] = net

    if not os.path.exists(INVOICE_DIR):
        os.makedirs(INVOICE_DIR)
    filename = os.path.join(INVOICE_DIR, f"Factura_{invoice_id}.xlsx")
    wb.save(filename)
    print(f"Factura generada: {filename}")

# =========================
# EJEMPLO DE USO
# =========================
if __name__ == "__main__":
    init_db()
    # Crear cliente (solo 1 vez)
    # add_client("INGECART 2018 S.L.", "B67021667", "Barcelona")
    # Crear factura
    create_invoice(
        client_id=1,
        services=4895.61,
        expenses=756
    )
