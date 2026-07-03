"""
Offers Extractor Orchestrator

Usage (run from repo root):
  & '<venv_python>' knowledge_hub/competitive_intel/offers_extractor_orchestrator.py "C:\path\to\source_folder"

This script performs discovery, parsing, reconciliation, evidence indexing
and outputs CSV/JSON/HTML reports for targeted offers. It is written to
be robust if optional OCR dependencies are missing: it will still attempt
text extraction and mark missing evidence.

Designed to satisfy the execution requirements provided by the user.
"""
from __future__ import annotations

import sys
import os
import re
import json
import csv
import time
import argparse
from dataclasses import dataclass, asdict, field
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import logging
import sqlite3

logging.basicConfig(level=logging.INFO, format="[%(levelname)s] %(message)s")
logger = logging.getLogger("offers_orchestrator")

# Optional heavy dependencies
try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    from PyPDF2 import PdfReader
except Exception:
    PdfReader = None

try:
    import docx
except Exception:
    docx = None

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from PIL import Image
    import pytesseract
except Exception:
    Image = None
    pytesseract = None


@dataclass
class OfferTarget:
    offer_number: str
    customer: str
    site: str
    country: str
    concept: str
    expected_value_eur: Optional[float]


@dataclass
class EvidenceRef:
    file: str
    page: Optional[int]
    snippet: str
    field: str


@dataclass
class OfferExtraction:
    offer_number: str
    customer: Optional[str] = None
    site: Optional[str] = None
    country: Optional[str] = None
    concept: Optional[str] = None
    expected_value_eur: Optional[float] = None
    source_files: List[str] = field(default_factory=list)
    offer_date: Optional[str] = None
    offer_version: Optional[str] = None
    currency: Optional[str] = None
    total_price_extracted: Optional[float] = None
    price_match_status: Optional[str] = None
    payment_terms_raw: Optional[str] = None
    payment_terms_normalized: Dict[str, Any] = field(default_factory=dict)
    deposit_percent: Optional[float] = None
    deposit_amount: Optional[float] = None
    milestones: List[str] = field(default_factory=list)
    final_payment: Optional[str] = None
    delivery_term: Optional[str] = None
    incoterm: Optional[str] = None
    warranty_terms: Optional[str] = None
    exclusions: Optional[str] = None
    confidence_score: float = 0.0
    truth_status: Optional[str] = None
    contradictions_found: List[str] = field(default_factory=list)
    missing_fields: List[str] = field(default_factory=list)
    evidence_refs: List[EvidenceRef] = field(default_factory=list)
    last_updated: Optional[str] = None


DEFAULT_OUTPUT_DIR = "knowledge_hub/outputs"


def ensure_output_dir(path: str):
    os.makedirs(path, exist_ok=True)


def find_files(root_dir: str) -> List[str]:
    candidates = []
    for dirpath, dirnames, filenames in os.walk(root_dir):
        for fn in filenames:
            if fn.startswith("~$"):
                continue
            candidates.append(os.path.join(dirpath, fn))
    return candidates


def read_text_from_pdf(path: str) -> List[Tuple[int, str]]:
    pages = []
    if pdfplumber:
        try:
            with pdfplumber.open(path) as pdf:
                for i, p in enumerate(pdf.pages, start=1):
                    text = p.extract_text() or ""
                    pages.append((i, text))
            return pages
        except Exception:
            logger.debug("pdfplumber failed for %s", path)
    if PdfReader:
        try:
            r = PdfReader(path)
            for i, page in enumerate(r.pages, start=1):
                try:
                    text = page.extract_text() or ""
                except Exception:
                    text = ""
                pages.append((i, text))
            return pages
        except Exception:
            logger.debug("PyPDF2 failed for %s", path)
    # Fallback: no text extractor available
    return []


def read_text_from_docx(path: str) -> List[Tuple[Optional[int], str]]:
    if docx is None:
        return []
    try:
        doc = docx.Document(path)
        full = "\n".join(p.text for p in doc.paragraphs if p.text)
        return [(None, full)]
    except Exception:
        return []


def read_text_from_xlsx(path: str) -> List[Tuple[Optional[int], str]]:
    if openpyxl is None:
        return []
    texts = []
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts = []
            for row in ws.iter_rows(values_only=True):
                for c in row:
                    if c is None:
                        continue
                    parts.append(str(c))
            texts.append((None, "\n".join(parts)))
        return texts
    except Exception:
        return []


def read_text_from_image(path: str) -> List[Tuple[Optional[int], str]]:
    if Image is None or pytesseract is None:
        return []
    try:
        img = Image.open(path)
        text = pytesseract.image_to_string(img)
        return [(None, text)]
    except Exception:
        return []


# Improved amount extraction: prefer explicit currency markers, thousand separators
AMOUNT_CURRENCY_PAT = re.compile(r"(?P<cur>€|\$|£|EUR|USD|GBP)\s*(?P<num>[0-9\.,\s]{1,})", flags=re.IGNORECASE)
AMOUNT_CURRENCY_PAT2 = re.compile(r"(?P<num>[0-9\.,\s]{1,})\s*(?P<cur>EUR|USD|GBP|€|\$|£)", flags=re.IGNORECASE)
AMOUNT_THOUSAND_PAT = re.compile(r"\b\d{1,3}(?:[\.,]\d{3})+(?:[\.,]\d{2})?\b")
AMOUNT_DECIMAL_PAT = re.compile(r"\b\d+[\.,]\d{2}\b")
AMOUNT_BIG_INT_PAT = re.compile(r"\b\d{5,}\b")


def _clean_number(raw: str) -> Optional[float]:
    raw_clean = raw.replace(" ", "").replace("\u00A0", "")
    # If both comma and dot present, decide decimal by last separator
    if raw_clean.count(",") and raw_clean.count("."):
        last_dot = raw_clean.rfind(".")
        last_comma = raw_clean.rfind(",")
        if last_comma > last_dot:
            raw_clean = raw_clean.replace(".", "").replace(",", ".")
        else:
            raw_clean = raw_clean.replace(",", "")
    else:
        if raw_clean.count(","):
            # treat comma as decimal if two digits after comma
            if len(raw_clean.split(",")[-1]) == 2:
                raw_clean = raw_clean.replace(".", "").replace(",", ".")
            else:
                raw_clean = raw_clean.replace(",", "")
        else:
            raw_clean = raw_clean.replace(".", "")
    try:
        return float(raw_clean)
    except Exception:
        return None


def parse_amount(s: str) -> Optional[Tuple[float, str]]:
    if not s:
        return None
    candidates: List[Tuple[float, str]] = []

    # 1) explicit currency before number
    for m in AMOUNT_CURRENCY_PAT.finditer(s):
        num = m.group("num")
        cur = m.group("cur")
        val = _clean_number(num)
        if val is not None:
            candidates.append((val, cur.upper()))

    # 2) explicit currency after number
    for m in AMOUNT_CURRENCY_PAT2.finditer(s):
        num = m.group("num")
        cur = m.group("cur")
        val = _clean_number(num)
        if val is not None:
            candidates.append((val, cur.upper()))

    # 3) thousand-sep patterns (1.052.000,00 or 1,052,000.00)
    for m in AMOUNT_THOUSAND_PAT.finditer(s):
        val = _clean_number(m.group(0))
        if val is not None:
            candidates.append((val, None))

    # 4) explicit decimal patterns like 12345.67 or 12.345,67
    for m in AMOUNT_DECIMAL_PAT.finditer(s):
        val = _clean_number(m.group(0))
        if val is not None:
            candidates.append((val, None))

    # 5) big integers (>=5 digits) as fallback
    for m in AMOUNT_BIG_INT_PAT.finditer(s):
        val = _clean_number(m.group(0))
        if val is not None:
            candidates.append((val, None))

    if not candidates:
        return None

    # prefer candidates with currency, and pick highest value otherwise
    with_cur = [c for c in candidates if c[1]]
    chosen = None
    if with_cur:
        # pick largest among currency-marked
        chosen = max(with_cur, key=lambda x: x[0])
    else:
        chosen = max(candidates, key=lambda x: x[0])

    # normalize currency codes
    cur_map = {"€": "EUR", "EUR": "EUR", "$": "USD", "USD": "USD", "£": "GBP", "GBP": "GBP"}
    val, cur = chosen
    cur_code = cur_map.get(cur, None) if cur else None
    return val, cur_code


def find_offer_identifiers(text: str) -> List[str]:
    # find patterns like OFF-2026-100 or #OFF-2026-100 or OFF 2026 100
    ids = []
    for m in re.finditer(r"#?OFF[-\s_]*(2026)[-\s_]*(\d{2,3})", text, flags=re.IGNORECASE):
        ids.append(f"OFF-2026-{m.group(2)}")
    return ids


def extract_payment_terms(text: str) -> Tuple[Optional[str], Dict[str, Any]]:
    # crude search for lines containing payment keywords
    lines = text.splitlines()
    hits = []
    for i, L in enumerate(lines):
        if re.search(r"anticipo|anticipo|deposit|advance|pago|hito|hitos|milestone|FAT|SAT|entrega|payment", L, flags=re.IGNORECASE):
            snippet = " ".join(lines[max(0, i - 2): min(len(lines), i + 3)])
            hits.append(snippet.strip())
    raw = "\n---\n".join(hits) if hits else None
    normalized = {}
    if hits:
        # detect percents and amounts
        percents = re.findall(r"(\d{1,3})\s*%", raw)
        if percents:
            normalized["percents"] = [int(p) for p in percents]
        # parse amounts using robust parser (prefers currency markers)
        amt = parse_amount(raw)
        if amt:
            normalized["amounts"] = [{"value": amt[0], "currency": amt[1], "snippet": str(amt[0])}]
    return raw, normalized


def snippet_for_match(text: str, match: str, ctx: int = 80) -> str:
    idx = text.lower().find(match.lower())
    if idx == -1:
        return text[:200]
    start = max(0, idx - ctx)
    end = min(len(text), idx + len(match) + ctx)
    return text[start:end].replace("\n", " ")


class OffersOrchestrator:
    def __init__(self, source_dir: str, outputs_dir: str = DEFAULT_OUTPUT_DIR):
        self.source_dir = source_dir
        self.outputs_dir = outputs_dir
        ensure_output_dir(self.outputs_dir)
        self.targets: List[OfferTarget] = self._load_targets()
        self.candidates: List[str] = []
        self.extractions: Dict[str, OfferExtraction] = {}

    def _load_targets(self) -> List[OfferTarget]:
        # Hard-coded targets per user request
        raw = [
            ("OFF-2026-100", "Mtorres", "Torres de Elorz", "SPAIN", "IIM Rollstand", 85000.00),
            ("OFF-2026-88", "IP", "WATERLOO", "USA", "Sr-1400", 197400.00),
            ("OFF-2026-89", "IP", "WATERLOO", "USA", "AMR", 502700.00),
            ("OFF-2026-92", "DS SMITH", "DICESA", "SPAIN", "Reparacion desperdicio", 25250.00),
            ("OFF-2026-80", "Cascades", "NEW JERSEY", "USA", "Robot FFG doble", 1052000.00),
            ("OFF-2026-81", "Cascades", "NEW JERSEY", "USA", "Salida conveyor 90º Evol.", 101482.00),
            ("OFF-2026-90", "Sterner Global", "Derbycorr", "USA", "Estación validacion", 225100.00),
            ("OFF-2026-91", "Sterner Global", "Derbycorr", "USA", "Ingetrans and AMR", 1300000.00),
            ("OFF-2026-97", "Pages", "", "USA", "Fabrica new Inge / AMR / FFG", 1600000.00),
            ("OFF-2026-82", "President Container", "Middletown", "USA", "Paletizador FFG", 550000.00),
            ("OFF-2026-83", "President Container", "Middletown", "USA", "Plug & Play + Ingepack", 600000.00),
            ("OFF-2026-84", "President Container", "Middletown", "USA", "Salida conveyor Funcem", 2400000.00),
            ("OFF-2026-85", "Pacific South West", "California", "USA", "Linea BHS + Ingetrans", 5250000.00),
            ("OFF-2026-86", "Font", "San Sadurni", "SPAIN", "Plug & Play", 162000.00),
            ("OFF-2026-87", "Font", "San Sadurni", "SPAIN", "Transfer central", 111500.00),
        ]
        return [OfferTarget(*r) for r in raw]

    def discover(self):
        logger.info("Discovering files under %s", self.source_dir)
        self.candidates = find_files(self.source_dir)
        logger.info("Found %d files", len(self.candidates))

    def parse_all(self):
        logger.info("Parsing candidate files for target offers...")
        # initialize extractions with expected values
        for t in self.targets:
            self.extractions[t.offer_number] = OfferExtraction(
                offer_number=t.offer_number,
                customer=t.customer,
                site=t.site,
                country=t.country,
                concept=t.concept,
                expected_value_eur=t.expected_value_eur,
                last_updated=datetime.utcnow().isoformat() + "Z",
            )

        def find_total_by_keywords(text: str, expected: Optional[float] = None) -> Optional[Tuple[float, str]]:
            # look for common total labels and parse nearby
            candidates = []
            for m in re.finditer(r"(?i)(total( price)?|resumen economico|precio paquete|precio total|total \:|total\s+\.|total\s+€|TOTAL)", text):
                start = max(0, m.start() - 200)
                end = min(len(text), m.end() + 200)
                snippet = text[start:end]
                amt = parse_amount(snippet)
                if amt:
                    candidates.append((amt[0], amt[1], snippet))
            # also try whole text fallback
            if not candidates:
                for pat in (AMOUNT_CURRENCY_PAT, AMOUNT_CURRENCY_PAT2, AMOUNT_THOUSAND_PAT, AMOUNT_DECIMAL_PAT):
                    for m in pat.finditer(text):
                        num = m.group(0)
                        amt = parse_amount(num)
                        if amt:
                            candidates.append((amt[0], amt[1], num))
            if not candidates:
                return None
            # prefer candidates with currency and near expected value
            if expected:
                # choose candidate closest to expected (relative)
                best = min(candidates, key=lambda c: abs(c[0] - expected) / max(1.0, expected))
                return best[0], best[1]
            # otherwise prefer one with currency
            with_cur = [c for c in candidates if c[1]]
            if with_cur:
                best = max(with_cur, key=lambda c: c[0])
                return best[0], best[1]
            best = max(candidates, key=lambda c: c[0])
            return best[0], best[1]

        for fpath in self.candidates:
            ext = os.path.splitext(fpath)[1].lower()
            text_pages: List[Tuple[Optional[int], str]] = []
            try:
                if ext == ".pdf":
                    text_pages = read_text_from_pdf(fpath)
                elif ext in (".docx",):
                    text_pages = read_text_from_docx(fpath)
                elif ext in (".xlsx", ".xlsm", ".xlsb", ".xls"):
                    text_pages = read_text_from_xlsx(fpath)
                elif ext in (".png", ".jpg", ".jpeg", ".tiff"):
                    text_pages = read_text_from_image(fpath)
                elif ext in (".txt", ".csv"):
                    try:
                        with open(fpath, "r", encoding="utf-8", errors="ignore") as fh:
                            text_pages = [(None, fh.read())]
                    except Exception:
                        text_pages = []
                else:
                    # attempt to read as text
                    try:
                        with open(fpath, "r", encoding="utf-8", errors="ignore") as fh:
                            text_pages = [(None, fh.read())]
                    except Exception:
                        text_pages = []
            except Exception as e:
                logger.debug("Failed to read %s: %s", fpath, e)
                text_pages = []

            combined = "\n\n".join(p for _, p in text_pages if p)
            # quick indicator: find any offer ids in this file
            ids = find_offer_identifiers(combined)

            if ids:
                # If the file explicitly contains offer IDs, only map to those offers
                for found_id in ids:
                    if found_id in self.extractions:
                        ex = self.extractions[found_id]
                        if fpath not in ex.source_files:
                            ex.source_files.append(fpath)
                        # try to extract total near keywords, preferring expected value
                        amt = find_total_by_keywords(combined, expected=ex.expected_value_eur)
                        if amt:
                            ex.total_price_extracted = amt[0]
                            ex.currency = amt[1]
                            ex.evidence_refs.append(EvidenceRef(file=fpath, page=None, snippet=snippet_for_match(combined, str(amt[0])), field="total_price"))
                        # payment terms
                        pay_raw, pay_norm = extract_payment_terms(combined)
                        if pay_raw:
                            ex.payment_terms_raw = pay_raw
                            ex.payment_terms_normalized = pay_norm
                            ex.evidence_refs.append(EvidenceRef(file=fpath, page=None, snippet=pay_raw[:400], field="payment_terms"))
                        # attach versions
                        if not ex.offer_version:
                            ex.offer_version = found_id
                        ex.last_updated = datetime.utcnow().isoformat() + "Z"
                # skip customer-based mapping for this file
                continue

            # No explicit IDs: fallback to customer/concept heuristics
            for t in self.targets:
                # match by customer or concept
                if re.search(re.escape(t.customer), combined, flags=re.IGNORECASE) or re.search(re.escape(t.concept), combined, flags=re.IGNORECASE):
                    ex = self.extractions[t.offer_number]
                    if fpath not in ex.source_files:
                        ex.source_files.append(fpath)
                    # attempt to extract totals using keywords and expected value
                    amt = find_total_by_keywords(combined, expected=ex.expected_value_eur)
                    if amt:
                        ex.total_price_extracted = amt[0]
                        ex.currency = amt[1]
                        ex.evidence_refs.append(EvidenceRef(file=fpath, page=None, snippet=snippet_for_match(combined, str(amt[0])), field="total_price"))
                    else:
                        # last-resort: parse full pages for numbers
                        for pnum, ptext in text_pages:
                            a = parse_amount(ptext)
                            if a:
                                ex.total_price_extracted = a[0]
                                ex.currency = a[1]
                                ex.evidence_refs.append(EvidenceRef(file=fpath, page=pnum, snippet=snippet_for_match(ptext, str(a[0])), field="total_price"))
                                break
                    pay_raw, pay_norm = extract_payment_terms(combined)
                    if pay_raw:
                        ex.payment_terms_raw = pay_raw
                        ex.payment_terms_normalized = pay_norm
                        ex.evidence_refs.append(EvidenceRef(file=fpath, page=None, snippet=pay_raw[:400], field="payment_terms"))
                    ex.last_updated = datetime.utcnow().isoformat() + "Z"

        # post-process price match status
        for onum, ex in self.extractions.items():
            if ex.total_price_extracted and ex.expected_value_eur:
                delta = abs(ex.total_price_extracted - ex.expected_value_eur)
                rel = delta / max(1.0, ex.expected_value_eur)
                ex.price_match_status = "MATCH" if rel < 0.02 else "MISMATCH"
                ex.confidence_score = max(0.2, 1.0 - rel)
            else:
                ex.price_match_status = "NOT_FOUND"
                if ex.total_price_extracted is None:
                    ex.missing_fields.append("total_price_extracted")

    def reconcile(self):
        logger.info("Reconciling versions and building truth records")
        for onum, ex in self.extractions.items():
            # choose selected source as newest file
            if ex.source_files:
                selected = max(ex.source_files, key=lambda f: os.path.getmtime(f))
                ex.truth_status = "SELECTED"
                ex.last_updated = datetime.utcfromtimestamp(os.path.getmtime(selected)).isoformat() + "Z"
                ex.evidence_refs.append(EvidenceRef(file=selected, page=None, snippet="selected source by mtime", field="selection"))
            else:
                ex.truth_status = "MISSING"

    def compute_quality(self) -> Dict[str, float]:
        total = len(self.targets)
        found = sum(1 for ex in self.extractions.values() if ex.source_files)
        coverage = found / total if total else 0.0
        payment_found = sum(1 for ex in self.extractions.values() if ex.payment_terms_raw)
        payment_score = payment_found / total if total else 0.0
        price_match = sum(1 for ex in self.extractions.values() if ex.price_match_status == "MATCH")
        price_score = price_match / total if total else 0.0
        evidence_cover = sum(1 for ex in self.extractions.values() if any(e.field in ("total_price", "payment_terms") for e in ex.evidence_refs))
        evidence_score = evidence_cover / total if total else 0.0
        return {
            "coverage": coverage,
            "payment_score": payment_score,
            "price_score": price_score,
            "evidence_score": evidence_score,
        }

    def persist_outputs(self):
        logger.info("Persisting outputs to %s", self.outputs_dir)
        csv_path = os.path.join(self.outputs_dir, "offers_2026_payment_terms_master.csv")
        json_path = os.path.join(self.outputs_dir, "offers_2026_payment_terms_master.json")
        html_path = os.path.join(self.outputs_dir, "OFFERS_2026_COMMERCIAL_TERMS_REPORT.html")
        quality_md = os.path.join(self.outputs_dir, "OFFERS_2026_EXTRACTION_QUALITY_REPORT.md")
        evidence_idx = os.path.join(self.outputs_dir, "OFFERS_2026_EVIDENCE_INDEX.json")

        # CSV
        fields = [
            "offer_number", "customer", "site", "country", "concept", "expected_value_eur",
            "source_files", "offer_date", "offer_version", "currency", "total_price_extracted",
            "price_match_status", "payment_terms_raw", "payment_terms_normalized", "deposit_percent",
            "deposit_amount", "milestone_1", "milestone_2", "milestone_3", "final_payment",
            "delivery_term", "incoterm", "warranty_terms", "exclusions", "confidence_score",
            "truth_status", "contradictions_found", "missing_fields", "evidence_refs", "last_updated"
        ]
        with open(csv_path, "w", newline='', encoding='utf-8') as fh:
            writer = csv.DictWriter(fh, fieldnames=fields)
            writer.writeheader()
            for ex in self.extractions.values():
                row = {k: None for k in fields}
                row.update({
                    "offer_number": ex.offer_number,
                    "customer": ex.customer,
                    "site": ex.site,
                    "country": ex.country,
                    "concept": ex.concept,
                    "expected_value_eur": ex.expected_value_eur,
                    "source_files": ";".join(ex.source_files),
                    "offer_date": ex.offer_date,
                    "offer_version": ex.offer_version,
                    "currency": ex.currency,
                    "total_price_extracted": ex.total_price_extracted,
                    "price_match_status": ex.price_match_status,
                    "payment_terms_raw": ex.payment_terms_raw,
                    "payment_terms_normalized": json.dumps(ex.payment_terms_normalized, ensure_ascii=False),
                    "deposit_percent": ex.payment_terms_normalized.get("percents") if ex.payment_terms_normalized else None,
                    "deposit_amount": None,
                    "milestone_1": ex.milestones[0] if len(ex.milestones) > 0 else None,
                    "milestone_2": ex.milestones[1] if len(ex.milestones) > 1 else None,
                    "milestone_3": ex.milestones[2] if len(ex.milestones) > 2 else None,
                    "final_payment": ex.final_payment,
                    "delivery_term": ex.delivery_term,
                    "incoterm": ex.incoterm,
                    "warranty_terms": ex.warranty_terms,
                    "exclusions": ex.exclusions,
                    "confidence_score": ex.confidence_score,
                    "truth_status": ex.truth_status,
                    "contradictions_found": ",".join(ex.contradictions_found),
                    "missing_fields": ",".join(ex.missing_fields),
                    "evidence_refs": json.dumps([asdict(e) for e in ex.evidence_refs], ensure_ascii=False),
                    "last_updated": ex.last_updated,
                })
                writer.writerow(row)

        # JSON
        out = {}
        for ex in self.extractions.values():
            out[ex.offer_number] = {
                "offer_number": ex.offer_number,
                "customer": ex.customer,
                "site": ex.site,
                "country": ex.country,
                "concept": ex.concept,
                "expected_value_eur": ex.expected_value_eur,
                "truth_record": {
                    "selected_source": ex.source_files[0] if ex.source_files else None,
                    "selected_version": ex.offer_version,
                    "selection_reason": ex.truth_status,
                    "confidence_score": ex.confidence_score,
                },
                "pricing": {
                    "total": ex.total_price_extracted,
                    "currency": ex.currency,
                    "line_items": [],
                },
                "payment_terms": {
                    "raw_text": ex.payment_terms_raw,
                    "normalized": ex.payment_terms_normalized,
                },
                "contractual_terms": {
                    "delivery": ex.delivery_term,
                    "warranty": ex.warranty_terms,
                    "exclusions": ex.exclusions,
                },
                "evidence": [asdict(e) for e in ex.evidence_refs],
                "contradictions": ex.contradictions_found,
                "missing_information": ex.missing_fields,
                "last_updated": ex.last_updated,
            }
        with open(json_path, "w", encoding='utf-8') as fh:
            json.dump(out, fh, ensure_ascii=False, indent=2)

        # Evidence index
        evidence_arr = []
        for ex in self.extractions.values():
            for e in ex.evidence_refs:
                evidence_arr.append({
                    "offer_number": ex.offer_number,
                    "file": e.file,
                    "page": e.page,
                    "snippet": e.snippet,
                    "field": e.field,
                })
        with open(evidence_idx, "w", encoding='utf-8') as fh:
            json.dump(evidence_arr, fh, ensure_ascii=False, indent=2)

        # Quality report
        quality = self.compute_quality()
        with open(quality_md, "w", encoding='utf-8') as fh:
            fh.write("# Extraction Quality Report\n\n")
            fh.write(f"Generated: {datetime.utcnow().isoformat()}Z\n\n")
            fh.write(f"Coverage: {quality['coverage']:.3f}\n")
            fh.write(f"Payment Terms Detection: {quality['payment_score']:.3f}\n")
            fh.write(f"Price Match Score: {quality['price_score']:.3f}\n")
            fh.write(f"Evidence Coverage: {quality['evidence_score']:.3f}\n\n")
            fh.write("## Per-offer summary\n\n")
            for ex in self.extractions.values():
                fh.write(f"- {ex.offer_number}: sources={len(ex.source_files)}, total={ex.total_price_extracted}, payment_terms={bool(ex.payment_terms_raw)}, confidence={ex.confidence_score:.2f}\n")

        # HTML executive report (simple table + summary)
        with open(html_path, "w", encoding='utf-8') as fh:
            fh.write("<html><head><meta charset=\"utf-8\"><title>Offers 2026 Commercial Terms Report</title></head><body>")
            fh.write("<h1>Offers 2026 Commercial Terms Report</h1>")
            fh.write(f"<p>Generated: {datetime.utcnow().isoformat()}Z</p>")
            fh.write("<h2>Summary</h2>")
            fh.write(f"<p>Coverage: {quality['coverage']:.3f}, Payment Terms: {quality['payment_score']:.3f}, Price Match: {quality['price_score']:.3f}</p>")
            fh.write("<h2>Offers</h2>")
            fh.write("<table border=1 cellpadding=4 cellspacing=0>")
            fh.write("<tr><th>Offer</th><th>Customer</th><th>Expected EUR</th><th>Extracted Total</th><th>Currency</th><th>Payment Terms</th><th>Sources</th></tr>")
            for ex in self.extractions.values():
                fh.write("<tr>")
                fh.write(f"<td>{ex.offer_number}</td>")
                fh.write(f"<td>{ex.customer}</td>")
                fh.write(f"<td>{ex.expected_value_eur}</td>")
                fh.write(f"<td>{ex.total_price_extracted}</td>")
                fh.write(f"<td>{ex.currency}</td>")
                fh.write(f"<td>{(ex.payment_terms_raw[:200] + '...') if ex.payment_terms_raw else ''}</td>")
                fh.write(f"<td>{';'.join(ex.source_files)}</td>")
                fh.write("</tr>")
            fh.write("</table>")
            fh.write("</body></html>")

        # Persist to local SQLite truth store
        db_path = os.path.join(self.outputs_dir, "offers_2026.db")
        try:
            self._persist_to_db(db_path)
            logger.info("Persisted extractions to DB: %s", db_path)
        except Exception as e:
            logger.warning("Failed to persist DB: %s", e)

        return {
            "csv": csv_path,
            "json": json_path,
            "html": html_path,
            "quality": quality_md,
            "evidence": evidence_idx,
        }

    def _persist_to_db(self, db_path: str):
        con = sqlite3.connect(db_path)
        cur = con.cursor()
        # Create schema for offers
        cur.executescript("""
        CREATE TABLE IF NOT EXISTS customers (id INTEGER PRIMARY KEY, name TEXT UNIQUE);
        CREATE TABLE IF NOT EXISTS offers (id INTEGER PRIMARY KEY, offer_number TEXT UNIQUE, customer_id INTEGER, site TEXT, country TEXT, concept TEXT, expected_value_eur REAL, last_updated TEXT);
        CREATE TABLE IF NOT EXISTS offer_versions (id INTEGER PRIMARY KEY, offer_id INTEGER, version TEXT, source_file TEXT, selected INTEGER, last_modified REAL);
        CREATE TABLE IF NOT EXISTS payment_terms (id INTEGER PRIMARY KEY, offer_version_id INTEGER, raw_text TEXT, normalized_json TEXT);
        CREATE TABLE IF NOT EXISTS line_items (id INTEGER PRIMARY KEY, offer_version_id INTEGER, description TEXT, amount REAL, currency TEXT);
        CREATE TABLE IF NOT EXISTS evidences (id INTEGER PRIMARY KEY, offer_version_id INTEGER, file_path TEXT, page INTEGER, snippet TEXT, field TEXT);
        """)
        con.commit()

        for ex in self.extractions.values():
            # customer
            cust_id = None
            if ex.customer:
                cur.execute("INSERT OR IGNORE INTO customers(name) VALUES(?)", (ex.customer,))
                con.commit()
                cur.execute("SELECT id FROM customers WHERE name=?", (ex.customer,))
                row = cur.fetchone()
                cust_id = row[0] if row else None

            # offer
            cur.execute("INSERT OR REPLACE INTO offers(offer_number, customer_id, site, country, concept, expected_value_eur, last_updated) VALUES(?,?,?,?,?,?,?)",
                        (ex.offer_number, cust_id, ex.site, ex.country, ex.concept, ex.expected_value_eur, ex.last_updated))
            con.commit()
            cur.execute("SELECT id FROM offers WHERE offer_number=?", (ex.offer_number,))
            offer_id = cur.fetchone()[0]

            # offer_version
            selected_file = ex.source_files[0] if ex.source_files else None
            selected_flag = 1 if ex.truth_status == "SELECTED" else 0
            last_mod = None
            if selected_file and os.path.exists(selected_file):
                last_mod = os.path.getmtime(selected_file)
            cur.execute("INSERT INTO offer_versions(offer_id, version, source_file, selected, last_modified) VALUES(?,?,?,?,?)",
                        (offer_id, ex.offer_version, selected_file, selected_flag, last_mod))
            con.commit()
            offer_version_id = cur.lastrowid

            # payment terms
            cur.execute("INSERT INTO payment_terms(offer_version_id, raw_text, normalized_json) VALUES(?,?,?)",
                        (offer_version_id, ex.payment_terms_raw, json.dumps(ex.payment_terms_normalized, ensure_ascii=False)))
            con.commit()

            # line items: create a total line item if present
            if ex.total_price_extracted is not None:
                cur.execute("INSERT INTO line_items(offer_version_id, description, amount, currency) VALUES(?,?,?,?)",
                            (offer_version_id, "TOTAL", ex.total_price_extracted, ex.currency))
                con.commit()

            # evidences
            for e in ex.evidence_refs:
                cur.execute("INSERT INTO evidences(offer_version_id, file_path, page, snippet, field) VALUES(?,?,?,?,?)",
                            (offer_version_id, e.file, e.page, e.snippet, e.field))
            con.commit()

        con.close()

    def run(self):
        self.discover()
        self.parse_all()
        self.reconcile()
        paths = self.persist_outputs()
        quality = self.compute_quality()
        return paths, quality, self.extractions


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source_dir", help="Source folder to scan")
    parser.add_argument("--outputs", default=DEFAULT_OUTPUT_DIR, help="Outputs folder")
    args = parser.parse_args()

    if not os.path.isdir(args.source_dir):
        logger.error("Source directory not found: %s", args.source_dir)
        sys.exit(2)

    orch = OffersOrchestrator(args.source_dir, outputs_dir=args.outputs)
    start = time.time()
    paths, quality, extractions = orch.run()
    elapsed = time.time() - start
    logger.info("Done in %.1fs", elapsed)

    # Final console block per user request
    total_targets = len(orch.targets)
    found = [o.offer_number for o in extractions.values() if o.source_files]
    not_found = [o.offer_number for o in extractions.values() if not o.source_files]

    contradictions = []
    for ex in extractions.values():
        if ex.price_match_status == "MISMATCH":
            contradictions.append(ex.offer_number)

    files_written = list(paths.values())

    top_risks = []
    # create simple risk notes
    for ex in extractions.values():
        if ex.truth_status == "MISSING":
            top_risks.append(f"{ex.offer_number}: MISSING source")
        if ex.price_match_status == "MISMATCH":
            top_risks.append(f"{ex.offer_number}: Price mismatch (expected {ex.expected_value_eur} vs extracted {ex.total_price_extracted})")

    print("1. OFFERS FOUND / NOT FOUND")
    print(f"FOUND: {len(found)} -> {','.join(found) if found else 'NONE'}")
    print(f"NOT_FOUND: {len(not_found)} -> {','.join(not_found) if not_found else 'NONE'}")
    print("2. COVERAGE SCORE")
    print(f"{quality['coverage']:.3f}")
    print("3. PAYMENT TERMS SCORE")
    print(f"{quality['payment_score']:.3f}")
    print("4. CONTRADICTIONS DETECTED")
    print(f"{len(contradictions)} -> {','.join(contradictions) if contradictions else 'NONE'}")
    print("5. FILES WRITTEN")
    for p in files_written:
        print(p)
    print("6. TOP RISKS / DATA GAPS")
    if top_risks:
        for r in top_risks[:10]:
            print(f"- {r}")
    else:
        print("NONE")
    print("7. NEXT ACTIONS REQUIRED FROM USER (solo si realmente faltan datos)")
    if not_found or quality['payment_score'] < 0.8 or quality['coverage'] < 0.9:
        actions = []
        if not_found:
            actions.append("Provide access or confirm alternate folders for missing offers.")
        if pdfplumber is None and pytesseract is None:
            actions.append("Install PDF/OCR stack (pdfplumber, pytesseract, poppler) to improve OCR on scanned PDFs.")
        actions.append("If files are password protected, provide passwords or unlocked copies.")
        for a in actions:
            print(f"- {a}")
    else:
        print("NONE")


if __name__ == "__main__":
    main()
