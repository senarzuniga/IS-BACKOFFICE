"""Report generator for Corrugated Plant Simulator.

Generates PDF (via ReportLab) and Excel (via openpyxl) reports
from SimulationResults objects.
"""
from __future__ import annotations

import io
import os
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

from .models import SimulationResults, MachineMetric, BottleneckRecord


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def generate_excel_report(results: SimulationResults) -> bytes:
    """Generate a formatted Excel workbook and return raw bytes."""
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.label import DataLabelList
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    wb = openpyxl.Workbook()

    # --- Styles ---
    ORANGE = "FF6A00"
    DARK = "0D1117"
    CARD = "1A1D24"
    TEXT = "F4F5F7"
    GREEN = "4FC17B"
    YELLOW = "F0C040"
    RED = "E05555"

    def hdr_fill(): return PatternFill("solid", fgColor=ORANGE)
    def dark_fill(): return PatternFill("solid", fgColor=CARD)
    def hdr_font(): return Font(bold=True, color=DARK, size=11)
    def body_font(): return Font(color=TEXT, size=10)
    def title_font(): return Font(bold=True, color=ORANGE, size=14)

    def write_row(ws, row, values, bold=False, fill_color=None):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(bold=bold, color=TEXT, size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if fill_color:
                cell.fill = PatternFill("solid", fgColor=fill_color)

    # =====================================================================
    # Sheet 1: Summary
    # =====================================================================
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_properties.tabColor = ORANGE
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 22

    ws1["A1"] = f"INGECART — Corrugated Plant Simulator"
    ws1["A1"].font = title_font()
    ws1["A2"] = f"Simulación: {results.plant_name}"
    ws1["A2"].font = Font(bold=True, color=TEXT, size=12)
    ws1["A3"] = f"Tipo: {results.plant_type} | Duración: {results.duration_hours:.0f}h | Generado: {results.completed_at}"
    ws1["A3"].font = Font(color="7E848E", size=9)
    ws1.row_dimensions[1].height = 22

    # KPI table
    kpis = [
        ("KPI", "Valor", "Referencia Sectorial"),
        ("m² producidos (total)", f"{results.m2_produced:,.0f}", "—"),
        ("Unidades convertidas", f"{results.total_units_converted:,.0f}", "—"),
        ("Eficiencia corrugadora", f"{results.corrugator_efficiency:.1f}%", "87-92%"),
        ("OEE medio planta", f"{results.average_oee:.1f}%", ">78% world class"),
        ("Utilización transporte", f"{results.transport_utilization:.1f}%", "<80% recomendado"),
        ("Buffer — ocupación media", f"{results.buffer_avg_occupancy:.1f}%", "30-60% óptimo"),
        ("Buffer — ocupación máxima", f"{results.buffer_max_occupancy:.1f}%", "<90%"),
        ("Escenario", results.scenario_label, "—"),
    ]
    for i, (label, val, ref) in enumerate(kpis):
        row = 5 + i
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=val)
        ws1.cell(row=row, column=3, value=ref)
        if i == 0:
            for col in range(1, 4):
                c = ws1.cell(row=row, column=col)
                c.fill = hdr_fill()
                c.font = hdr_font()
        else:
            for col in range(1, 4):
                c = ws1.cell(row=row, column=col)
                c.font = body_font()
                c.fill = dark_fill()

    # Recommendations
    ws1["A15"] = "RECOMENDACIONES AUTOMÁTICAS"
    ws1["A15"].font = Font(bold=True, color=ORANGE, size=11)
    for i, rec in enumerate(results.recommendations[:6]):
        ws1[f"A{16+i}"] = rec.replace("**", "").replace("_", "")
        ws1[f"A{16+i}"].font = body_font()
        ws1[f"A{16+i}"].fill = dark_fill()
        ws1.row_dimensions[16+i].height = 22

    ws1.sheet_view.showGridLines = False

    # =====================================================================
    # Sheet 2: Machine Metrics
    # =====================================================================
    ws2 = wb.create_sheet("Métricas Máquinas")
    ws2.sheet_properties.tabColor = "4FC17B"
    headers = ["Máquina", "Disponibilidad %", "Rendimiento %", "Calidad %", "OEE %",
               "Producción", "Tiempo Bloqueado (s)", "Tiempo Setup (s)"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill(); c.font = hdr_font()
        ws2.column_dimensions[c.column_letter].width = 18

    for row, mm in enumerate(results.machine_metrics, 2):
        vals = [
            mm.machine_id, mm.availability, mm.performance,
            mm.quality, mm.oee, round(mm.units_produced, 0),
            round(mm.blocked_time_s, 0), round(mm.setup_time_s, 0)
        ]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=row, column=col, value=val)
            c.font = body_font(); c.fill = dark_fill()
            if col in (2, 3, 4, 5):
                oee_val = val if isinstance(val, (int, float)) else 0
                if oee_val >= 85:
                    c.fill = PatternFill("solid", fgColor="1A3D1A")
                elif oee_val >= 75:
                    c.fill = PatternFill("solid", fgColor="3D3000")
                else:
                    c.fill = PatternFill("solid", fgColor="3D1A1A")

    ws2.sheet_view.showGridLines = False

    # =====================================================================
    # Sheet 3: Timeline
    # =====================================================================
    if results.timeline:
        ws3 = wb.create_sheet("Timeline (1 min)")
        ws3.sheet_properties.tabColor = "4A90D9"
        if results.timeline:
            headers3 = list(results.timeline[0].keys())
            for col, h in enumerate(headers3, 1):
                c = ws3.cell(row=1, column=col, value=h)
                c.fill = hdr_fill(); c.font = hdr_font()
                ws3.column_dimensions[c.column_letter].width = 16
            for row, snap in enumerate(results.timeline, 2):
                for col, h in enumerate(headers3, 1):
                    c = ws3.cell(row=row, column=col, value=snap.get(h))
                    c.font = body_font(); c.fill = dark_fill()
        ws3.sheet_view.showGridLines = False

    # =====================================================================
    # Sheet 4: Bottlenecks
    # =====================================================================
    ws4 = wb.create_sheet("Cuellos de Botella")
    ws4.sheet_properties.tabColor = "E05555"
    b_headers = ["Ubicación", "Tipo", "Espera Media (s)", "Espera Máxima (s)", "Frecuencia"]
    for col, h in enumerate(b_headers, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.fill = hdr_fill(); c.font = hdr_font()
        ws4.column_dimensions[c.column_letter].width = 20

    for row, bt in enumerate(results.bottlenecks, 2):
        vals = [bt.location, bt.type, bt.avg_wait_s, bt.max_wait_s, bt.frequency]
        for col, val in enumerate(vals, 1):
            c = ws4.cell(row=row, column=col, value=val)
            c.font = body_font()
            c.fill = PatternFill("solid", fgColor="3D1A1A")

    if not results.bottlenecks:
        ws4["A2"] = "✅ No se detectaron cuellos de botella significativos"
        ws4["A2"].font = Font(color=GREEN, size=11, bold=True)

    ws4.sheet_view.showGridLines = False

    # Save and return bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF report
# ---------------------------------------------------------------------------

def generate_pdf_report(results: SimulationResults) -> bytes:
    """Generate a styled PDF report and return raw bytes."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        raise RuntimeError("reportlab is required: pip install reportlab")

    ORANGE = colors.HexColor("#FF6A00")
    DARK = colors.HexColor("#0D1117")
    CARD = colors.HexColor("#1A1D24")
    TEXT = colors.HexColor("#F4F5F7")
    GREEN = colors.HexColor("#4FC17B")
    YELLOW = colors.HexColor("#F0C040")
    RED = colors.HexColor("#E05555")
    MUTED = colors.HexColor("#7E848E")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"],
        textColor=ORANGE, fontSize=22, spaceAfter=4, fontName="Helvetica-Bold",
        backColor=DARK,
    )
    h2_style = ParagraphStyle(
        "H2", parent=styles["Heading2"],
        textColor=ORANGE, fontSize=13, fontName="Helvetica-Bold",
        spaceBefore=14, spaceAfter=4,
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        textColor=TEXT, fontSize=9, fontName="Helvetica",
        spaceAfter=4, leading=13,
    )
    muted_style = ParagraphStyle(
        "Muted", parent=styles["Normal"],
        textColor=MUTED, fontSize=8, fontName="Helvetica",
    )

    story = []

    # Title
    story.append(Paragraph("INGECART — Corrugated Plant Simulator", title_style))
    story.append(Paragraph(f"<b>{results.plant_name}</b>", h2_style))
    story.append(Paragraph(
        f"Tipo: {results.plant_type} | Duración: {results.duration_hours:.0f}h | "
        f"Escenario: {results.scenario_label} | Generado: {results.completed_at}",
        muted_style,
    ))
    story.append(HRFlowable(color=ORANGE, thickness=1, spaceAfter=10))

    # KPI Summary table
    story.append(Paragraph("Resultados Principales", h2_style))
    kpi_data = [
        ["KPI", "Resultado", "Referencia Sectorial"],
        ["m² producidos (total)", f"{results.m2_produced:,.0f}", "—"],
        ["Unidades convertidas", f"{results.total_units_converted:,.0f}", "—"],
        ["Eficiencia corrugadora", f"{results.corrugator_efficiency:.1f}%", "87–92%"],
        ["OEE medio planta", f"{results.average_oee:.1f}%", ">78% (world class)"],
        ["Utilización transporte", f"{results.transport_utilization:.1f}%", "<80%"],
        ["Buffer — ocupación media", f"{results.buffer_avg_occupancy:.1f}%", "30–60%"],
        ["Buffer — ocupación máxima", f"{results.buffer_max_occupancy:.1f}%", "<90%"],
    ]
    tbl = Table(kpi_data, colWidths=[7*cm, 4*cm, 5*cm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), DARK),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), CARD),
        ("TEXTCOLOR", (0, 1), (-1, -1), TEXT),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [CARD, colors.HexColor("#21252e")]),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#333")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 12))

    # Machine metrics
    story.append(Paragraph("Métricas por Máquina", h2_style))
    mm_data = [["Máquina", "OEE %", "Disponib. %", "Rendim. %", "Calidad %", "Producción"]]
    for mm in results.machine_metrics:
        mm_data.append([
            mm.machine_id,
            f"{mm.oee:.1f}",
            f"{mm.availability:.1f}",
            f"{mm.performance:.1f}",
            f"{mm.quality:.1f}",
            f"{mm.units_produced:,.0f}",
        ])
    mm_tbl = Table(mm_data, colWidths=[3.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 3*cm])
    mm_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
        ("TEXTCOLOR", (0, 0), (-1, 0), DARK),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 1), (-1, -1), CARD),
        ("TEXTCOLOR", (0, 1), (-1, -1), TEXT),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#333")),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    mm_tbl.setStyle(TableStyle(mm_style_cmds))
    story.append(mm_tbl)
    story.append(Spacer(1, 12))

    # Bottlenecks
    story.append(Paragraph("Cuellos de Botella Detectados", h2_style))
    if results.bottlenecks:
        bt_data = [["Ubicación", "Tipo", "Espera Media", "Espera Máx.", "Frecuencia"]]
        for bt in results.bottlenecks[:6]:
            bt_data.append([
                bt.location, bt.type,
                f"{bt.avg_wait_s:.0f}s", f"{bt.max_wait_s:.0f}s", str(bt.frequency),
            ])
        bt_tbl = Table(bt_data, colWidths=[4.5*cm, 3*cm, 3*cm, 3*cm, 3*cm])
        bt_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ORANGE),
            ("TEXTCOLOR", (0, 0), (-1, 0), DARK),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#2d0a0a")),
            ("TEXTCOLOR", (0, 1), (-1, -1), TEXT),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#444")),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ]))
        story.append(bt_tbl)
    else:
        story.append(Paragraph("✅ No se detectaron cuellos de botella significativos.", body_style))
    story.append(Spacer(1, 12))

    # Recommendations
    story.append(Paragraph("Recomendaciones de Optimización", h2_style))
    for rec in results.recommendations:
        clean = rec.replace("**", "").replace("__", "")
        story.append(Paragraph(f"• {clean}", body_style))

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(color=ORANGE, thickness=0.5))
    story.append(Paragraph(
        "INGECART — Corrugated Plant Simulator | Ing_RESEARCH | "
        f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        muted_style,
    ))

    doc.build(story)
    return buf.getvalue()
